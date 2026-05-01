"""Convert hpi_at_zip5.xlsx to a compact zip->year->hpi JSON."""
import json
from openpyxl import load_workbook

SRC = "C:/Users/user1/Documents/house/hpi_at_zip5.xlsx"
DST = "C:/Users/user1/Documents/house/hpi_at_zip5.json"

wb = load_workbook(SRC, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

rows = ws.iter_rows(values_only=True)
header_idx = -1
col_zip = col_year = col_hpi = -1
buffered = []
for i, row in enumerate(rows):
    buffered.append(row)
    norm = [str(c).lower().strip() if c is not None else "" for c in row]
    zip_idx = next((j for j, c in enumerate(norm) if "zip" in c or c == "fips"), -1)
    yr_idx = next((j for j, c in enumerate(norm) if c == "year" or c == "yr"), -1)
    hpi_idx = next((j for j, c in enumerate(norm) if c == "hpi"), -1)
    if zip_idx != -1 and yr_idx != -1 and hpi_idx != -1:
        header_idx = i
        col_zip, col_year, col_hpi = zip_idx, yr_idx, hpi_idx
        break
    if i >= 9:
        break

if header_idx == -1:
    raise SystemExit("Header row not found")

data = {}
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i <= header_idx:
        continue
    if not row or row[col_zip] is None or row[col_year] is None or row[col_hpi] is None:
        continue
    try:
        zip_str = str(row[col_zip]).strip().zfill(5)
        year = int(row[col_year])
        hpi = float(row[col_hpi])
    except (ValueError, TypeError):
        continue
    if hpi <= 0:
        continue
    data.setdefault(zip_str, {})[str(year)] = round(hpi, 2)

with open(DST, "w") as f:
    json.dump(data, f, separators=(",", ":"))

import os
print(f"Wrote {DST}: {len(data)} ZIPs, {os.path.getsize(DST)/1_000_000:.2f} MB")
